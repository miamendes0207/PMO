# modules/action_utils.py

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .action_config import ACTION_SHEET_NAME, ACTION_COLUMNS

DATA_ROOT = Path("data/actions")
DATA_ROOT.mkdir(parents=True, exist_ok=True)

TEMPLATE_ACTIONS_PATH = DATA_ROOT / "Template_Actions.xlsx"


def get_client_actions_file(client_name: str) -> Path:
    safe = client_name.replace(" ", "_")
    return DATA_ROOT / f"{safe}_Actions.xlsx"


def _create_template_if_missing() -> None:
    """Create a blank template actions workbook if it doesn't exist."""
    if TEMPLATE_ACTIONS_PATH.exists():
        return

    df = pd.DataFrame(columns=ACTION_COLUMNS)
    with pd.ExcelWriter(TEMPLATE_ACTIONS_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=ACTION_SHEET_NAME, index=False)


def _ensure_client_file(client_name: str) -> Path:
    """Ensure the client actions workbook exists (copy from template if needed)."""
    _create_template_if_missing()
    client_file = get_client_actions_file(client_name)

    if not client_file.exists():
        client_file.write_bytes(TEMPLATE_ACTIONS_PATH.read_bytes())

    return client_file


def load_client_actions(client_name: str) -> pd.DataFrame:
    """Load the actions sheet for a client."""
    client_file = _ensure_client_file(client_name)
    try:
        df = pd.read_excel(client_file, sheet_name=ACTION_SHEET_NAME)
    except ValueError:
        # Sheet missing – initialise it
        df = pd.DataFrame(columns=ACTION_COLUMNS)
        save_client_actions(client_name, df)
    return df


def save_client_actions(client_name: str, df: pd.DataFrame) -> None:
    """
    Save the actions DataFrame back to the client workbook,
    preserving header row where possible.
    """
    client_file = _ensure_client_file(client_name)
    wb = load_workbook(client_file)

    if ACTION_SHEET_NAME in wb.sheetnames:
        ws = wb[ACTION_SHEET_NAME]
    else:
        ws = wb.create_sheet(ACTION_SHEET_NAME)

    # Ensure header row exists
    if ws.max_row == 0:
        for col_idx, col_name in enumerate(ACTION_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
    else:
        # If header row exists but differs, we won't overwrite styles
        pass

    # Clear all data rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write DataFrame rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row[col_name])

    wb.save(client_file)
