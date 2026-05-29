# modules/gov_log_utils.py

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook

LOG_SHEET_NAME = "Log"

LOG_COLUMNS = [
    "Log ID",
    "Client",
    "Period Start",
    "Period End",
    "Generated At",
    "File Name",
    "Sections",
    "Notes",
]

DATA_ROOT = Path("data/governance")
DATA_ROOT.mkdir(parents=True, exist_ok=True)

TEMPLATE_LOG_PATH = DATA_ROOT / "Template_Gov_Log.xlsx"


def _create_template_if_missing():
    if TEMPLATE_LOG_PATH.exists():
        return
    df = pd.DataFrame(columns=LOG_COLUMNS)
    with pd.ExcelWriter(TEMPLATE_LOG_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=LOG_SHEET_NAME, index=False)


def get_client_log_file(client: str) -> Path:
    safe = client.replace(" ", "_")
    return DATA_ROOT / f"{safe}_Gov_Log.xlsx"


def _ensure_client_log_file(client: str) -> Path:
    _create_template_if_missing()
    client_file = get_client_log_file(client)
    if not client_file.exists():
        client_file.write_bytes(TEMPLATE_LOG_PATH.read_bytes())
    return client_file


def load_client_log(client: str) -> pd.DataFrame:
    client_file = _ensure_client_log_file(client)
    try:
        df = pd.read_excel(client_file, sheet_name=LOG_SHEET_NAME)
    except ValueError:
        df = pd.DataFrame(columns=LOG_COLUMNS)
        save_client_log(client, df)
    return df


def save_client_log(client: str, df: pd.DataFrame) -> None:
    client_file = _ensure_client_log_file(client)
    wb = load_workbook(client_file)

    if LOG_SHEET_NAME in wb.sheetnames:
        ws = wb[LOG_SHEET_NAME]
    else:
        ws = wb.create_sheet(LOG_SHEET_NAME)

    # Ensure header
    if ws.max_row == 0:
        for col_idx, col_name in enumerate(LOG_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

    # Clear data rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write df
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row[col_name])

    wb.save(client_file)


def append_log_entry(
    client: str,
    period_start,
    period_end,
    generated_at,
    file_name: str,
    sections: List[str],
    notes: str = "",
) -> int:
    df = load_client_log(client)
    if df.empty:
        next_id = 1
    else:
        max_id = pd.to_numeric(df["Log ID"], errors="coerce").max()
        next_id = int(max_id) + 1 if pd.notna(max_id) else 1

    new_row = {
        "Log ID": next_id,
        "Client": client,
        "Period Start": period_start,
        "Period End": period_end,
        "Generated At": generated_at,
        "File Name": file_name,
        "Sections": ", ".join(sections),
        "Notes": notes,
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    save_client_log(client, df)
    return next_id
